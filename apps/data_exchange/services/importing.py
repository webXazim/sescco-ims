from __future__ import annotations

import re
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from apps.inventory.models import StockItem, Unit
from apps.inventory.normalization import clean_display_text, normalize_phone, normalize_text
from apps.inventory.services.matching import find_stock_matches
from apps.inventory.services.stock import InventoryOperationError, add_opening_stock
from apps.projects.models import Project

from ..models import ImportJob, ImportRow
from .workbook_security import WorkbookArchiveError, validate_workbook_archive


class ImportPreviewError(ValueError):
    pass


class ImportProcessingError(ValueError):
    pass


_HEADER_CLEANER = re.compile(r"[^a-z0-9]+")


LEGACY_ALIASES = {
    "material_name": {"materialname", "material"},
    "description": {"description", "specification"},
    "supplier_name": {"nameofsupplier", "suppliername", "supplier"},
    "supplier_phone": {
        "contactmobileno",
        "contactmobilenumber",
        "mobilecontactno",
        "mobileno",
        "mobile",
        "phone",
        "supplierphone",
    },
    "supplier_location": {"locationofsupplier", "supplierlocation", "location"},
    "unit_price": {"unitprice", "unitpricerate", "price", "rate"},
    "purchase_date": {"lastpurchasedate", "purchasedate", "date"},
}

OPENING_ALIASES = {
    "project_code": {"projectcode", "project"},
    "material_name": {"materialname", "material"},
    "description": {"description", "specification"},
    "supplier_name": {"suppliername", "nameofsupplier", "supplier"},
    "supplier_phone": {"supplierphone", "contactmobileno", "phone", "mobile"},
    "supplier_location": {"supplierlocation", "locationofsupplier", "location"},
    "unit": {"unit", "unitsymbol", "unitname"},
    "opening_quantity": {"openingquantity", "quantity", "openingstock"},
    "unit_price": {"unitprice", "price", "rate"},
    "opening_date": {"openingdate", "date"},
    "minimum_quantity": {"minimumquantity", "minimum", "reorderlevel"},
    "reference": {"reference", "invoicereference", "invoice"},
    "notes": {"notes", "note"},
}


@dataclass(frozen=True)
class ParsedPhone:
    display: str
    warning: str = ""


def _header_key(value) -> str:
    return _HEADER_CLEANER.sub("", normalize_text(str(value or "")))


def _display(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_display_text(str(value))


def _json_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _phone(value) -> ParsedPhone:
    if value is None:
        return ParsedPhone("")
    warning = ""
    if isinstance(value, bool):
        return ParsedPhone("")
    if isinstance(value, (int, float)):
        display = str(int(value)) if float(value).is_integer() else str(value)
        warning = (
            "Phone was stored as a number in Excel and may have lost a leading zero or + sign."
        )
    else:
        display = clean_display_text(str(value))
    return ParsedPhone(display=display, warning=warning)


def _decimal(value, *, required: bool = False) -> Decimal | None:
    if value in (None, ""):
        if required:
            raise ValueError("A numeric value is required.")
        return None
    if isinstance(value, str):
        value = value.replace(",", "").strip()
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError(f"Invalid numeric value: {value}") from exc


def _date(value, *, workbook, required: bool = False) -> date | None:
    if value in (None, ""):
        if required:
            raise ValueError("A date is required.")
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value, workbook.epoch)
        return converted.date() if isinstance(converted, datetime) else converted
    text = clean_display_text(str(value))
    for pattern in (
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m/%d/%Y",
        "%d %b %Y",
        "%d %B %Y",
    ):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date value: {text}")


def _open_workbook(job: ImportJob):
    try:
        job.source_file.open("rb")
        try:
            payload = job.source_file.read()
        finally:
            job.source_file.close()
    except Exception as exc:
        raise ImportPreviewError("The uploaded workbook could not be read.") from exc

    try:
        validate_workbook_archive(payload)
    except WorkbookArchiveError as exc:
        raise ImportPreviewError(str(exc)) from exc

    try:
        return load_workbook(
            BytesIO(payload),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise ImportPreviewError("The workbook could not be opened as XLSX/XLSM.") from exc


def _select_sheet(workbook, preferred: str):
    for sheet_name in workbook.sheetnames:
        if normalize_text(sheet_name) == normalize_text(preferred):
            return workbook[sheet_name]
    return workbook[workbook.sheetnames[0]]


def _find_headers(worksheet, aliases: dict[str, set[str]], required: set[str]):
    max_scan = min(worksheet.max_row or 1, 30)
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_scan, values_only=True),
        start=1,
    ):
        mapping: dict[str, int] = {}
        for index, value in enumerate(values):
            key = _header_key(value)
            if not key:
                continue
            for field, field_aliases in aliases.items():
                if field not in mapping and key in field_aliases:
                    mapping[field] = index
        if required.issubset(mapping):
            return row_number, mapping
    required_text = ", ".join(sorted(required))
    raise ImportPreviewError(f"Could not find the required columns: {required_text}.")


def _row_dict(values, mapping: dict[str, int]) -> dict[str, object]:
    result = {}
    for field, index in mapping.items():
        result[field] = values[index] if index < len(values) else None
    return result


def _message(parts: list[str]) -> str:
    return " ".join(part.strip() for part in parts if part and part.strip())


def _create_preview_row(
    *,
    job: ImportJob,
    row_number: int,
    status: str,
    action: str,
    raw: dict,
    cleaned: dict,
    messages: list[str],
    exact_match: StockItem | None = None,
    similar_ids: list[int] | None = None,
    requires_confirmation: bool = False,
) -> ImportRow:
    return ImportRow.objects.create(
        job=job,
        row_number=row_number,
        status=status,
        planned_action=action,
        requires_confirmation=requires_confirmation,
        raw_data={key: _json_value(value) for key, value in raw.items()},
        cleaned_data={key: _json_value(value) for key, value in cleaned.items()},
        message=_message(messages),
        exact_match=exact_match,
        similar_match_ids=similar_ids or [],
    )


def _preview_legacy(job: ImportJob, workbook) -> None:
    if not job.project_id or not job.default_unit_id:
        raise ImportPreviewError("Legacy imports require a project and default unit.")
    worksheet = _select_sheet(workbook, "Database")
    header_row, mapping = _find_headers(
        worksheet,
        LEGACY_ALIASES,
        {"material_name", "supplier_name", "supplier_phone"},
    )
    seen: dict[tuple[str, str, str], int] = {}
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        raw = _row_dict(values, mapping)
        if not any(value not in (None, "") for value in raw.values()):
            continue
        messages: list[str] = []
        try:
            parsed_phone = _phone(raw.get("supplier_phone"))
            if parsed_phone.warning:
                messages.append(parsed_phone.warning)
            cleaned = {
                "material_name": _display(raw.get("material_name")),
                "description": _display(raw.get("description")),
                "supplier_name": _display(raw.get("supplier_name")),
                "supplier_phone": parsed_phone.display,
                "supplier_location": _display(raw.get("supplier_location")),
                "unit_price": _decimal(raw.get("unit_price")),
                "purchase_date": _date(raw.get("purchase_date"), workbook=workbook),
            }
            if not cleaned["material_name"]:
                raise ValueError("Material name is required.")
            if not cleaned["supplier_name"]:
                raise ValueError("Supplier name is required.")
            if not normalize_phone(cleaned["supplier_phone"]):
                raise ValueError("Supplier phone is required.")
            if cleaned["unit_price"] is not None and cleaned["unit_price"] < 0:
                raise ValueError("Unit price cannot be negative.")
            identity = (
                normalize_text(cleaned["material_name"]),
                normalize_text(cleaned["supplier_name"]),
                normalize_phone(cleaned["supplier_phone"]),
            )
            if identity in seen:
                _create_preview_row(
                    job=job,
                    row_number=row_number,
                    status=ImportRow.Status.WARNING,
                    action=ImportRow.Action.SKIP,
                    raw=raw,
                    cleaned=cleaned,
                    messages=[f"Duplicate of workbook row {seen[identity]}; it will be skipped."],
                )
                continue
            seen[identity] = row_number
            matches = find_stock_matches(
                project=job.project,
                material_name=cleaned["material_name"],
                supplier_name=cleaned["supplier_name"],
                supplier_phone=cleaned["supplier_phone"],
            )
            similar = list(matches.similar)
            if matches.exact:
                action = ImportRow.Action.UPDATE
                status = ImportRow.Status.WARNING if messages else ImportRow.Status.VALID
                messages.append("An exact stock record exists and can be updated.")
                _create_preview_row(
                    job=job,
                    row_number=row_number,
                    status=status,
                    action=action,
                    raw=raw,
                    cleaned=cleaned,
                    messages=messages,
                    exact_match=matches.exact,
                )
            elif similar:
                messages.append(
                    "Project, material and supplier match an existing record, "
                    "but the phone differs."
                )
                _create_preview_row(
                    job=job,
                    row_number=row_number,
                    status=ImportRow.Status.WARNING,
                    action=ImportRow.Action.CREATE_SEPARATE,
                    raw=raw,
                    cleaned=cleaned,
                    messages=messages,
                    similar_ids=[item.pk for item in similar],
                    requires_confirmation=True,
                )
            else:
                _create_preview_row(
                    job=job,
                    row_number=row_number,
                    status=ImportRow.Status.WARNING if messages else ImportRow.Status.VALID,
                    action=ImportRow.Action.CREATE,
                    raw=raw,
                    cleaned=cleaned,
                    messages=messages or ["A new zero-balance stock record will be created."],
                )
        except ValueError as exc:
            _create_preview_row(
                job=job,
                row_number=row_number,
                status=ImportRow.Status.ERROR,
                action=ImportRow.Action.SKIP,
                raw=raw,
                cleaned={},
                messages=[str(exc)],
            )


def _preview_opening(job: ImportJob, workbook) -> None:
    worksheet = _select_sheet(workbook, "Opening Stock")
    header_row, mapping = _find_headers(
        worksheet,
        OPENING_ALIASES,
        {
            "project_code",
            "material_name",
            "supplier_name",
            "supplier_phone",
            "unit",
            "opening_quantity",
            "opening_date",
        },
    )
    seen: dict[tuple[int, str, str, str], int] = {}
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        raw = _row_dict(values, mapping)
        if not any(value not in (None, "") for value in raw.values()):
            continue
        messages: list[str] = []
        try:
            parsed_phone = _phone(raw.get("supplier_phone"))
            if parsed_phone.warning:
                messages.append(parsed_phone.warning)
            project_code = _display(raw.get("project_code")).upper()
            unit_value = _display(raw.get("unit"))
            project = (
                Project.objects.filter(deleted_at__isnull=True)
                .filter(
                    code=project_code,
                    status=Project.Status.ACTIVE,
                )
                .first()
            )
            if not project:
                raise ValueError(f"Active project {project_code or '—'} was not found.")
            normalized_unit = normalize_text(unit_value)
            unit = (
                Unit.objects.filter(is_active=True, deleted_at__isnull=True)
                .filter(Q(normalized_name=normalized_unit) | Q(normalized_symbol=normalized_unit))
                .first()
            )
            if not unit:
                raise ValueError(f"Active unit {unit_value or '—'} was not found.")
            cleaned = {
                "project_id": project.pk,
                "project_code": project.code,
                "material_name": _display(raw.get("material_name")),
                "description": _display(raw.get("description")),
                "supplier_name": _display(raw.get("supplier_name")),
                "supplier_phone": parsed_phone.display,
                "supplier_location": _display(raw.get("supplier_location")),
                "unit_id": unit.pk,
                "unit": unit.symbol,
                "opening_quantity": _decimal(raw.get("opening_quantity"), required=True),
                "unit_price": _decimal(raw.get("unit_price")),
                "opening_date": _date(raw.get("opening_date"), workbook=workbook, required=True),
                "minimum_quantity": _decimal(raw.get("minimum_quantity")) or Decimal("0"),
                "minimum_quantity_provided": raw.get("minimum_quantity") not in (None, ""),
                "reference": _display(raw.get("reference")),
                "notes": _display(raw.get("notes")),
            }
            if not cleaned["material_name"]:
                raise ValueError("Material name is required.")
            if not cleaned["supplier_name"]:
                raise ValueError("Supplier name is required.")
            if not normalize_phone(cleaned["supplier_phone"]):
                raise ValueError("Supplier phone is required.")
            if cleaned["opening_quantity"] <= 0:
                raise ValueError("Opening quantity must be greater than zero.")
            if cleaned["minimum_quantity"] < 0:
                raise ValueError("Minimum quantity cannot be negative.")
            if cleaned["unit_price"] is not None and cleaned["unit_price"] < 0:
                raise ValueError("Unit price cannot be negative.")
            if cleaned["opening_date"] > timezone.localdate():
                raise ValueError("Opening date cannot be in the future.")
            identity = (
                project.pk,
                normalize_text(cleaned["material_name"]),
                normalize_text(cleaned["supplier_name"]),
                normalize_phone(cleaned["supplier_phone"]),
            )
            if identity in seen:
                _create_preview_row(
                    job=job,
                    row_number=row_number,
                    status=ImportRow.Status.ERROR,
                    action=ImportRow.Action.SKIP,
                    raw=raw,
                    cleaned=cleaned,
                    messages=[f"Duplicate opening identity in workbook row {seen[identity]}."],
                )
                continue
            seen[identity] = row_number
            matches = find_stock_matches(
                project=project,
                material_name=cleaned["material_name"],
                supplier_name=cleaned["supplier_name"],
                supplier_phone=cleaned["supplier_phone"],
            )
            similar = list(matches.similar)
            if matches.exact:
                if matches.exact.status != StockItem.Status.ACTIVE:
                    raise ValueError(
                        "Matching stock record is archived and cannot receive opening stock."
                    )
                if matches.exact.unit_id != unit.pk:
                    raise ValueError(
                        f"Matching stock uses {matches.exact.unit.symbol}, not {unit.symbol}."
                    )
                if matches.exact.current_quantity != 0 or matches.exact.movements.exists():
                    raise ValueError(
                        "Matching stock already has quantity or movement history; opening stock "
                        "cannot be applied."
                    )
                _create_preview_row(
                    job=job,
                    row_number=row_number,
                    status=ImportRow.Status.WARNING if messages else ImportRow.Status.VALID,
                    action=ImportRow.Action.OPEN_EXISTING,
                    raw=raw,
                    cleaned=cleaned,
                    messages=(
                        messages or ["Opening stock will be added to the empty matching record."]
                    ),
                    exact_match=matches.exact,
                )
            elif similar:
                messages.append(
                    "Project, material and supplier match an existing record, "
                    "but the phone differs."
                )
                _create_preview_row(
                    job=job,
                    row_number=row_number,
                    status=ImportRow.Status.WARNING,
                    action=ImportRow.Action.CREATE_SEPARATE,
                    raw=raw,
                    cleaned=cleaned,
                    messages=messages,
                    similar_ids=[item.pk for item in similar],
                    requires_confirmation=True,
                )
            else:
                _create_preview_row(
                    job=job,
                    row_number=row_number,
                    status=ImportRow.Status.WARNING if messages else ImportRow.Status.VALID,
                    action=ImportRow.Action.CREATE_OPENING,
                    raw=raw,
                    cleaned=cleaned,
                    messages=(
                        messages or ["A new stock record and opening movement will be created."]
                    ),
                )
        except ValueError as exc:
            _create_preview_row(
                job=job,
                row_number=row_number,
                status=ImportRow.Status.ERROR,
                action=ImportRow.Action.SKIP,
                raw=raw,
                cleaned={},
                messages=[str(exc)],
            )


def _refresh_counts(job: ImportJob) -> None:
    rows = job.rows.all()
    job.total_rows = rows.count()
    job.valid_rows = rows.filter(status=ImportRow.Status.VALID).count()
    job.warning_rows = rows.filter(status=ImportRow.Status.WARNING).count()
    job.error_rows = rows.filter(status=ImportRow.Status.ERROR).count()
    job.save(
        update_fields=(
            "total_rows",
            "valid_rows",
            "warning_rows",
            "error_rows",
        )
    )


def preview_import(job: ImportJob) -> ImportJob:
    workbook = None
    try:
        workbook = _open_workbook(job)
        with transaction.atomic():
            job.rows.all().delete()
            job.status = ImportJob.Status.PREVIEW
            job.error_message = ""
            job.save(update_fields=("status", "error_message"))
            if job.import_type == ImportJob.Type.LEGACY_CATALOG:
                _preview_legacy(job, workbook)
            elif job.import_type == ImportJob.Type.OPENING_STOCK:
                _preview_opening(job, workbook)
            else:
                raise ImportPreviewError("Unsupported import type.")
            _refresh_counts(job)
    except Exception as exc:
        job.status = ImportJob.Status.FAILED
        job.error_message = str(exc)
        job.save(update_fields=("status", "error_message"))
        if isinstance(exc, ImportPreviewError):
            raise
        raise ImportPreviewError(str(exc)) from exc
    finally:
        if workbook is not None:
            workbook.close()
    return job


def _legacy_create(row: ImportRow, job: ImportJob, user) -> StockItem:
    data = row.cleaned_data
    item = StockItem(
        project=job.project,
        material_name=data["material_name"],
        description=data.get("description", ""),
        supplier_name=data["supplier_name"],
        supplier_phone=data["supplier_phone"],
        supplier_location=data.get("supplier_location", ""),
        unit=job.default_unit,
        current_quantity=Decimal("0"),
        latest_unit_price=Decimal(data["unit_price"]) if data.get("unit_price") else None,
        latest_addition_date=(
            date.fromisoformat(data["purchase_date"]) if data.get("purchase_date") else None
        ),
        created_by=user,
        updated_by=user,
    )
    item.save(_inventory_service=True)
    return item


def _legacy_update(row: ImportRow, job: ImportJob, user) -> StockItem:
    item = StockItem.objects.select_for_update().get(pk=row.exact_match_id)
    if not job.options.get("update_existing_records", True):
        raise ImportProcessingError("Matching-row updates were disabled for this job.")
    data = row.cleaned_data
    update_fields = ["updated_by", "updated_at"]
    if data.get("description"):
        item.description = data["description"]
        update_fields.append("description")
    if data.get("supplier_location"):
        item.supplier_location = data["supplier_location"]
        update_fields.append("supplier_location")
    imported_date = date.fromisoformat(data["purchase_date"]) if data.get("purchase_date") else None
    imported_price = Decimal(data["unit_price"]) if data.get("unit_price") else None
    if imported_date and (
        not item.latest_addition_date or imported_date >= item.latest_addition_date
    ):
        item.latest_addition_date = imported_date
        update_fields.append("latest_addition_date")
        if imported_price is not None:
            item.latest_unit_price = imported_price
            update_fields.append("latest_unit_price")
    elif item.latest_unit_price is None and imported_price is not None:
        item.latest_unit_price = imported_price
        update_fields.append("latest_unit_price")
    item.updated_by = user
    item.save(_inventory_service=True, update_fields=tuple(dict.fromkeys(update_fields)))
    return item


def _opening_item(row: ImportRow, user) -> StockItem:
    data = row.cleaned_data
    if row.exact_match_id:
        item = StockItem.objects.select_for_update().get(pk=row.exact_match_id)
        update_fields = ["updated_by", "updated_at"]
        if data.get("description"):
            item.description = data["description"]
            update_fields.append("description")
        if data.get("supplier_location"):
            item.supplier_location = data["supplier_location"]
            update_fields.append("supplier_location")
        if data.get("minimum_quantity_provided"):
            item.minimum_quantity = Decimal(data["minimum_quantity"])
            update_fields.append("minimum_quantity")
        item.updated_by = user
        item.save(update_fields=tuple(dict.fromkeys(update_fields)))
        return item
    item = StockItem(
        project_id=data["project_id"],
        material_name=data["material_name"],
        description=data.get("description", ""),
        supplier_name=data["supplier_name"],
        supplier_phone=data["supplier_phone"],
        supplier_location=data.get("supplier_location", ""),
        unit_id=data["unit_id"],
        minimum_quantity=Decimal(data.get("minimum_quantity") or "0"),
        created_by=user,
        updated_by=user,
    )
    item.save()
    return item


def confirm_import(
    *,
    job: ImportJob,
    user,
    include_similar_rows: bool = False,
) -> ImportJob:
    if not getattr(user, "is_inventory_admin", False):
        raise ImportProcessingError("Only an administrator can confirm imports.")
    try:
        with transaction.atomic():
            locked_job = ImportJob.objects.select_for_update().get(pk=job.pk)
            if locked_job.status == ImportJob.Status.COMPLETED:
                return locked_job
            if locked_job.status != ImportJob.Status.PREVIEW:
                raise ImportProcessingError("Only a previewed import can be confirmed.")
            locked_job.status = ImportJob.Status.PROCESSING
            locked_job.save(update_fields=("status",))
            imported = 0
            skipped = 0
            rows = locked_job.rows.select_for_update().order_by("row_number")
            for row in rows:
                if (
                    row.status == ImportRow.Status.ERROR
                    or row.planned_action == ImportRow.Action.SKIP
                ):
                    row.status = ImportRow.Status.SKIPPED
                    row.processed_at = timezone.now()
                    row.save(update_fields=("status", "processed_at"))
                    skipped += 1
                    continue
                if row.requires_confirmation and not include_similar_rows:
                    row.status = ImportRow.Status.SKIPPED
                    row.processed_at = timezone.now()
                    row.message = _message(
                        [row.message, "Skipped because similar rows were not approved."]
                    )
                    row.save(update_fields=("status", "processed_at", "message"))
                    skipped += 1
                    continue
                if (
                    locked_job.import_type == ImportJob.Type.LEGACY_CATALOG
                    and row.planned_action == ImportRow.Action.UPDATE
                    and not locked_job.options.get("update_existing_records", True)
                ):
                    row.status = ImportRow.Status.SKIPPED
                    row.processed_at = timezone.now()
                    row.message = _message(
                        [row.message, "Skipped because matching updates were disabled."]
                    )
                    row.save(update_fields=("status", "processed_at", "message"))
                    skipped += 1
                    continue

                if locked_job.import_type == ImportJob.Type.LEGACY_CATALOG:
                    if row.planned_action == ImportRow.Action.UPDATE:
                        item = _legacy_update(row, locked_job, user)
                    else:
                        item = _legacy_create(row, locked_job, user)
                    movement = None
                elif locked_job.import_type == ImportJob.Type.OPENING_STOCK:
                    item = _opening_item(row, user)
                    data = row.cleaned_data
                    token = uuid.uuid5(
                        uuid.NAMESPACE_URL,
                        f"opening-import:{locked_job.reference}:{row.row_number}",
                    )
                    try:
                        result = add_opening_stock(
                            stock_item=item,
                            user=user,
                            idempotency_key=token,
                            quantity=Decimal(data["opening_quantity"]),
                            movement_date=date.fromisoformat(data["opening_date"]),
                            unit_price=(
                                Decimal(data["unit_price"]) if data.get("unit_price") else None
                            ),
                            invoice_reference=data.get("reference", ""),
                            notes=data.get("notes", ""),
                        )
                    except InventoryOperationError as exc:
                        raise ImportProcessingError(str(exc)) from exc
                    movement = result.movement
                else:
                    raise ImportProcessingError("Unsupported import type.")

                row.status = ImportRow.Status.IMPORTED
                row.imported_stock_item = item
                row.movement = movement
                row.processed_at = timezone.now()
                row.save(
                    update_fields=(
                        "status",
                        "imported_stock_item",
                        "movement",
                        "processed_at",
                    )
                )
                imported += 1

            locked_job.status = ImportJob.Status.COMPLETED
            locked_job.imported_rows = imported
            locked_job.skipped_rows = skipped
            locked_job.confirmed_at = timezone.now()
            locked_job.error_message = ""
            locked_job.save(
                update_fields=(
                    "status",
                    "imported_rows",
                    "skipped_rows",
                    "confirmed_at",
                    "error_message",
                )
            )
            return locked_job
    except Exception as exc:
        ImportJob.objects.filter(pk=job.pk).update(
            status=ImportJob.Status.FAILED,
            error_message=str(exc),
        )
        if isinstance(exc, ImportProcessingError):
            raise
        raise ImportProcessingError(str(exc)) from exc
