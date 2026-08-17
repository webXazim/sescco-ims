from __future__ import annotations

import csv
import re
from collections.abc import Callable, Iterable
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO

from django.db.models import F
from django.http import HttpResponse
from django.http.request import QueryDict
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from apps.explorer.filtering import querydict_to_plain, resolve_date_range
from apps.inventory.forms import MovementFilterForm, StockHistoryFilterForm, StockItemFilterForm
from apps.inventory.models import StockItem, StockMovement, Unit
from apps.inventory.selectors import (
    apply_movement_search,
    apply_stock_search,
    apply_stock_status,
    filter_stock_items,
    filter_stock_movements,
    low_stock_items,
    stock_items,
    stock_movements,
)
from apps.projects.models import Project

from ..models import ExportAudit
from ..opening_schema import OPENING_IMPORT_COLUMNS, OPENING_IMPORT_RULES

DEFAULT_STOCK_COLUMNS = (
    "project",
    "condition",
    "material",
    "supplier",
    "phone",
    "quantity",
    "minimum",
    "unit",
    "price",
    "latest_addition",
    "stock_status",
    "updated",
)
DEFAULT_MOVEMENT_COLUMNS = (
    "date",
    "project",
    "condition",
    "material",
    "type",
    "quantity",
    "balance",
    "reference",
    "user",
)
STOCK_SORTS = {
    "project": ("location__code", "material_name", "supplier_name"),
    "material": ("material_name", "supplier_name"),
    "-material": ("-material_name", "supplier_name"),
    "supplier": ("supplier_name", "material_name"),
    "-supplier": ("-supplier_name", "material_name"),
    "quantity": ("current_quantity", "material_name"),
    "-quantity": ("-current_quantity", "material_name"),
    "minimum": ("minimum_quantity", "material_name"),
    "-minimum": ("-minimum_quantity", "material_name"),
    "updated": ("-updated_at", "material_name"),
    "created": ("-created_at", "material_name"),
    "latest-addition": (F("latest_addition_date").desc(nulls_last=True), "material_name"),
    "oldest-addition": (F("latest_addition_date").asc(nulls_last=True), "material_name"),
    "price": (F("latest_unit_price").asc(nulls_last=True), "material_name"),
    "-price": (F("latest_unit_price").desc(nulls_last=True), "material_name"),
}
MOVEMENT_SORTS = {
    "-date": ("-movement_date", "-created_at", "-pk"),
    "date": ("movement_date", "created_at", "pk"),
    "project": ("location_code_snapshot", "material_name_snapshot", "-movement_date"),
    "material": ("material_name_snapshot", "-movement_date"),
    "type": ("movement_type", "-movement_date"),
    "quantity": ("quantity", "-movement_date"),
    "-quantity": ("-quantity", "-movement_date"),
    "user": ("created_by__first_name", "created_by__username", "-movement_date"),
}


@dataclass(frozen=True)
class ExportColumn:
    key: str
    label: str
    value: Callable[[object], object]
    width: int = 18
    number_format: str | None = None


@dataclass(frozen=True)
class ExportDataset:
    audit_dataset: str
    title: str
    queryset: object
    columns: tuple[ExportColumn, ...]
    filters: dict
    sort: str
    scope_reference: str = ""
    scope_label: str = ""


FORMULA_PREFIXES = ("=", "+", "-", "@")
_FILENAME_CLEANER = re.compile(r"[^a-zA-Z0-9._-]+")


def safe_text(value: object) -> object:
    if not isinstance(value, str):
        return value
    cleaned = value.replace("\x00", "")
    if cleaned.lstrip().startswith(FORMULA_PREFIXES):
        return "'" + cleaned
    return cleaned


def _decimal_value(value: Decimal | None):
    return float(value) if value is not None else None


def _user_name(user) -> str:
    return user.display_name if user else ""


def _reference(movement: StockMovement) -> str:
    return next(
        (
            value
            for value in (
                movement.invoice_reference,
                movement.purpose,
                movement.recipient,
                movement.reason,
            )
            if value
        ),
        "",
    )


STOCK_COLUMNS = {
    "project": ExportColumn(
        "project",
        "Project",
        lambda item: f"{item.location.code} · {item.location.name}",
        28,
    ),
    "condition": ExportColumn(
        "condition", "Condition", lambda item: item.get_condition_display(), 16
    ),
    "material": ExportColumn("material", "Material", lambda item: item.material_name, 26),
    "description": ExportColumn("description", "Description", lambda item: item.description, 36),
    "supplier": ExportColumn("supplier", "Supplier", lambda item: item.supplier_name, 26),
    "phone": ExportColumn("phone", "Supplier phone", lambda item: item.supplier_phone, 20),
    "location": ExportColumn(
        "location",
        "Supplier location",
        lambda item: item.supplier_location,
        24,
    ),
    "quantity": ExportColumn(
        "quantity",
        "Quantity",
        lambda item: _decimal_value(item.current_quantity),
        14,
        "0.000",
    ),
    "minimum": ExportColumn(
        "minimum",
        "Minimum quantity",
        lambda item: _decimal_value(item.minimum_quantity),
        16,
        "0.000",
    ),
    "unit": ExportColumn("unit", "Unit", lambda item: item.unit.symbol, 12),
    "price": ExportColumn(
        "price",
        "Latest unit price",
        lambda item: _decimal_value(item.latest_unit_price),
        16,
        "0.00",
    ),
    "latest_addition": ExportColumn(
        "latest_addition",
        "Latest addition",
        lambda item: item.latest_addition_date,
        16,
        "yyyy-mm-dd",
    ),
    "stock_status": ExportColumn(
        "stock_status",
        "Stock status",
        lambda item: (
            "Archived" if item.status == StockItem.Status.ARCHIVED else item.stock_status_label
        ),
        16,
    ),
    "updated": ExportColumn(
        "updated",
        "Last updated",
        lambda item: item.updated_at,
        20,
        "yyyy-mm-dd hh:mm",
    ),
    "created": ExportColumn(
        "created",
        "Created",
        lambda item: item.created_at,
        20,
        "yyyy-mm-dd hh:mm",
    ),
}

MOVEMENT_COLUMNS = {
    "date": ExportColumn(
        "date",
        "Movement date",
        lambda movement: movement.movement_date,
        16,
        "yyyy-mm-dd",
    ),
    "project": ExportColumn(
        "project",
        "Location",
        lambda movement: (f"{movement.location_code_display} · {movement.location_name_display}"),
        28,
    ),
    "condition": ExportColumn(
        "condition",
        "Condition",
        lambda movement: movement.condition_display.replace("_", " ").title(),
        16,
    ),
    "material": ExportColumn(
        "material",
        "Material",
        lambda movement: movement.material_name_display,
        26,
    ),
    "supplier": ExportColumn(
        "supplier",
        "Supplier",
        lambda movement: movement.supplier_name_display,
        26,
    ),
    "phone": ExportColumn(
        "phone",
        "Supplier phone",
        lambda movement: movement.supplier_phone_display,
        20,
    ),
    "type": ExportColumn(
        "type",
        "Action",
        lambda movement: movement.get_movement_type_display(),
        20,
    ),
    "quantity": ExportColumn(
        "quantity",
        "Signed quantity",
        lambda movement: _decimal_value(movement.signed_quantity),
        16,
        "0.000",
    ),
    "balance": ExportColumn(
        "balance",
        "Balance",
        lambda movement: f"{movement.previous_balance} → {movement.new_balance}",
        22,
    ),
    "price": ExportColumn(
        "price",
        "Unit price",
        lambda movement: _decimal_value(movement.unit_price),
        14,
        "0.00",
    ),
    "reference": ExportColumn(
        "reference",
        "Reference / purpose",
        _reference,
        32,
    ),
    "user": ExportColumn(
        "user",
        "User",
        lambda movement: _user_name(movement.created_by),
        20,
    ),
}


def _with_defaults(query: QueryDict, defaults: dict[str, object]) -> QueryDict:
    data = query.copy()
    data.pop("page", None)
    data.pop("movement_page", None)
    for key, value in defaults.items():
        if key in data:
            continue
        if isinstance(value, tuple | list):
            data.setlist(key, [str(item) for item in value])
        else:
            data[key] = str(value)
    return data


def _visible_columns(
    query: QueryDict,
    *,
    choices: Iterable[tuple[str, str]],
    defaults: tuple[str, ...],
    mapping: dict[str, ExportColumn],
) -> tuple[ExportColumn, ...]:
    allowed = {key for key, _ in choices}
    requested = [key for key in query.getlist("columns") if key in allowed]
    selected = requested or list(defaults)
    return tuple(mapping[key] for key in selected if key in mapping)


def inventory_dataset(query: QueryDict, *, low_stock: bool = False) -> ExportDataset:
    data = _with_defaults(
        query,
        {
            "status": StockItem.Status.ACTIVE,
            "date_field": "latest_addition_date",
            "sort": "quantity" if low_stock else "project",
            "columns": DEFAULT_STOCK_COLUMNS,
        },
    )
    form = StockItemFilterForm(data)
    if not form.is_valid():
        raise ValueError("The current inventory filters are invalid.")
    cleaned = form.cleaned_data
    queryset = low_stock_items() if low_stock else stock_items()
    if low_stock:
        scoped = dict(cleaned)
        scoped["status"] = StockItem.Status.ACTIVE
        if scoped.get("stock_status") not in {"low", "out"}:
            scoped["stock_status"] = ""
        queryset = filter_stock_items(queryset, scoped)
        cleaned = scoped
    else:
        queryset = filter_stock_items(queryset, cleaned)
    sort = cleaned.get("sort") or ("quantity" if low_stock else "project")
    queryset = queryset.order_by(*STOCK_SORTS.get(sort, STOCK_SORTS["project"]))
    columns = _visible_columns(
        data,
        choices=StockItemFilterForm.COLUMN_CHOICES,
        defaults=DEFAULT_STOCK_COLUMNS,
        mapping=STOCK_COLUMNS,
    )
    return ExportDataset(
        audit_dataset=(
            ExportAudit.Dataset.LOW_STOCK if low_stock else ExportAudit.Dataset.INVENTORY
        ),
        title="Low stock" if low_stock else "Inventory",
        queryset=queryset,
        columns=columns,
        filters=querydict_to_plain(data, form.fields.keys()),
        sort=sort,
    )


def activity_dataset(query: QueryDict) -> ExportDataset:
    data = _with_defaults(query, {"sort": "-date", "columns": DEFAULT_MOVEMENT_COLUMNS})
    form = MovementFilterForm(data)
    if not form.is_valid():
        raise ValueError("The current activity filters are invalid.")
    cleaned = form.cleaned_data
    queryset = filter_stock_movements(stock_movements(), cleaned)
    sort = cleaned.get("sort") or "-date"
    queryset = queryset.order_by(*MOVEMENT_SORTS.get(sort, MOVEMENT_SORTS["-date"]))
    columns = _visible_columns(
        data,
        choices=MovementFilterForm.COLUMN_CHOICES,
        defaults=DEFAULT_MOVEMENT_COLUMNS,
        mapping=MOVEMENT_COLUMNS,
    )
    return ExportDataset(
        audit_dataset=ExportAudit.Dataset.ACTIVITY,
        title="Stock activity",
        queryset=queryset,
        columns=columns,
        filters=querydict_to_plain(data, form.fields.keys()),
        sort=sort,
    )


def stock_history_dataset(stock_item: StockItem, query: QueryDict) -> ExportDataset:
    data = _with_defaults(query, {"sort": "-date"})
    form = StockHistoryFilterForm(data)
    if not form.is_valid():
        raise ValueError("The current stock-history filters are invalid.")
    cleaned = form.cleaned_data
    queryset = stock_movements().filter(stock_item=stock_item)
    queryset = apply_movement_search(queryset, cleaned.get("q") or "")
    if cleaned.get("movement_type"):
        queryset = queryset.filter(movement_type__in=cleaned["movement_type"])
    if cleaned.get("created_by"):
        queryset = queryset.filter(created_by=cleaned["created_by"])
    date_range = resolve_date_range(
        cleaned.get("date_preset") or "",
        start=cleaned.get("date_from"),
        end=cleaned.get("date_to"),
    )
    if date_range.start:
        queryset = queryset.filter(movement_date__gte=date_range.start)
    if date_range.end:
        queryset = queryset.filter(movement_date__lte=date_range.end)
    sort = cleaned.get("sort") or "-date"
    queryset = queryset.order_by(*MOVEMENT_SORTS.get(sort, MOVEMENT_SORTS["-date"]))
    columns = tuple(MOVEMENT_COLUMNS[key] for key in DEFAULT_MOVEMENT_COLUMNS)
    return ExportDataset(
        audit_dataset=ExportAudit.Dataset.STOCK_HISTORY,
        title=f"Stock history · {stock_item.material_name}",
        queryset=queryset,
        columns=columns,
        filters=querydict_to_plain(data, form.fields.keys()),
        sort=sort,
        scope_reference=str(stock_item.reference),
        scope_label=f"{stock_item.location.code} · {stock_item.material_name}",
    )


def project_inventory_dataset(project: Project, query: QueryDict) -> ExportDataset:
    queryset = stock_items().filter(project=project)
    search = query.get("q", "").strip()
    stock_status = query.get("stock_status", "").strip()
    record_status = query.get("record_status", StockItem.Status.ACTIVE).strip()
    if record_status == StockItem.Status.ARCHIVED:
        queryset = queryset.filter(status=StockItem.Status.ARCHIVED)
    elif record_status != "all":
        record_status = StockItem.Status.ACTIVE
        queryset = queryset.filter(status=StockItem.Status.ACTIVE)
    queryset = apply_stock_search(queryset, search)
    queryset = apply_stock_status(queryset, stock_status)
    queryset = queryset.order_by("material_name", "supplier_name")
    columns = tuple(
        STOCK_COLUMNS[key]
        for key in (
            "material",
            "description",
            "supplier",
            "phone",
            "quantity",
            "minimum",
            "unit",
            "price",
            "stock_status",
            "updated",
        )
    )
    filters = {
        "q": search,
        "stock_status": stock_status,
        "record_status": record_status,
    }
    return ExportDataset(
        audit_dataset=ExportAudit.Dataset.PROJECT_INVENTORY,
        title=f"Project inventory · {project.code}",
        queryset=queryset,
        columns=columns,
        filters={key: value for key, value in filters.items() if value},
        sort="material",
        scope_reference=project.code,
        scope_label=project.name,
    )


def _cell_value(value):
    value = safe_text(value)
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime) and timezone.is_aware(value):
        return timezone.localtime(value).replace(tzinfo=None)
    return value


def _metadata_rows(dataset: ExportDataset, user, row_count: int) -> list[tuple[str, object]]:
    rows: list[tuple[str, object]] = [
        ("Dataset", dataset.title),
        ("Scope", dataset.scope_label or "All matching records"),
        ("Exported by", user.display_name),
        ("Exported at", timezone.localtime().strftime("%Y-%m-%d %H:%M:%S %Z")),
        ("Matching rows", row_count),
        ("Sort", dataset.sort),
        ("Columns", ", ".join(column.label for column in dataset.columns)),
    ]
    for key, value in dataset.filters.items():
        if value in (None, "", []):
            continue
        display = ", ".join(str(item) for item in value) if isinstance(value, list) else value
        rows.append((f"Filter · {key}", display))
    return rows


def _xlsx_payload(dataset: ExportDataset, user, row_count: int) -> bytes:
    workbook = Workbook(write_only=True)
    results = workbook.create_sheet("Results")
    results.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="172033")
    header_font = Font(color="FFFFFF", bold=True)
    header_cells = []
    for column in dataset.columns:
        cell = WriteOnlyCell(results, value=column.label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        header_cells.append(cell)
    results.append(header_cells)

    for item in dataset.queryset.iterator(chunk_size=1000):
        row = []
        for column in dataset.columns:
            cell = WriteOnlyCell(results, value=_cell_value(column.value(item)))
            if column.number_format:
                cell.number_format = column.number_format
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            row.append(cell)
        results.append(row)
    results.auto_filter.ref = f"A1:{_excel_column(len(dataset.columns))}{row_count + 1}"
    for index, column in enumerate(dataset.columns, start=1):
        results.column_dimensions[_excel_column(index)].width = column.width

    info = workbook.create_sheet("Export Information")
    info.append(["Export information", "Value"])
    for label, value in _metadata_rows(dataset, user, row_count):
        info.append([label, _cell_value(value)])
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 72

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _csv_payload(dataset: ExportDataset) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow([column.label for column in dataset.columns])
    for item in dataset.queryset.iterator(chunk_size=1000):
        values = []
        for column in dataset.columns:
            value = _cell_value(column.value(item))
            if isinstance(value, date | datetime):
                value = (
                    value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
                )
            values.append(value)
        writer.writerow(values)
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _excel_column(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _filename(dataset: ExportDataset, file_format: str) -> str:
    date_part = timezone.localdate().isoformat()
    base = _FILENAME_CLEANER.sub("-", dataset.title.lower()).strip("-")
    return f"{base}-{date_part}.{file_format}"


def export_response(
    *,
    dataset: ExportDataset,
    user,
    file_format: str,
) -> HttpResponse:
    if file_format not in ExportAudit.Format.values:
        raise ValueError("Unsupported export format.")
    row_count = dataset.queryset.count()
    if file_format == ExportAudit.Format.XLSX:
        payload = _xlsx_payload(dataset, user, row_count)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        payload = _csv_payload(dataset)
        content_type = "text/csv; charset=utf-8"
    ExportAudit.objects.create(
        dataset=dataset.audit_dataset,
        file_format=file_format,
        filters=dataset.filters,
        columns=[column.key for column in dataset.columns],
        sort=dataset.sort,
        scope_reference=dataset.scope_reference,
        scope_label=dataset.scope_label,
        row_count=row_count,
        created_by=user,
    )
    response = HttpResponse(payload, content_type=content_type)
    response["Content-Disposition"] = f'attachment; filename="{_filename(dataset, file_format)}"'
    response["Cache-Control"] = "private, no-store"
    return response


def opening_stock_template_response(*, user) -> HttpResponse:
    projects = list(
        Project.objects.filter(status=Project.Status.ACTIVE, deleted_at__isnull=True).order_by(
            "code"
        )
    )
    units = list(Unit.objects.filter(is_active=True, deleted_at__isnull=True).order_by("name"))
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions.append(["Opening stock import template"])
    for index, rule in enumerate(OPENING_IMPORT_RULES, start=1):
        instructions.append([str(index), rule])
    instructions.append([])
    instructions.append(["Accepted columns", "Required", "Format", "Example"])
    for column in OPENING_IMPORT_COLUMNS:
        instructions.append(
            [
                column["name"],
                "Required" if column["required"] else "Optional",
                column["format"],
                column["example"],
            ]
        )
    instructions.column_dimensions["A"].width = 8
    instructions.column_dimensions["B"].width = 22
    instructions.column_dimensions["C"].width = 36
    instructions.column_dimensions["D"].width = 28
    instructions["A1"].font = Font(size=16, bold=True)
    for cell in instructions[8]:
        cell.fill = PatternFill("solid", fgColor="172033")
        cell.font = Font(color="FFFFFF", bold=True)
    for row in instructions.iter_rows(min_row=9, max_row=8 + len(OPENING_IMPORT_COLUMNS)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet = workbook.create_sheet("Opening Stock")
    headers = [column["name"] for column in OPENING_IMPORT_COLUMNS]
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:M1000"
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="172033")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    widths = [18, 26, 32, 26, 20, 24, 12, 18, 14, 16, 18, 20, 36]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[_excel_column(index)].width = width
    for row in range(2, 1001):
        sheet.cell(row=row, column=1).number_format = "@"
        sheet.cell(row=row, column=5).number_format = "@"
        sheet.cell(row=row, column=8).number_format = "0.000"
        sheet.cell(row=row, column=9).number_format = "0.00"
        sheet.cell(row=row, column=10).number_format = "yyyy-mm-dd"
        sheet.cell(row=row, column=11).number_format = "0.000"
        sheet.cell(row=row, column=12).number_format = "@"
    sheet.conditional_formatting.add(
        "H2:H1000",
        FormulaRule(formula=["H2<=0"], fill=PatternFill("solid", fgColor="FEE2E2")),
    )

    lists = workbook.create_sheet("Lists")
    lists.append(["Project Codes", "Units"])
    for index, project in enumerate(projects, start=2):
        lists.cell(row=index, column=1, value=project.code)
    for index, unit in enumerate(units, start=2):
        lists.cell(row=index, column=2, value=unit.symbol)
    project_end = max(2, len(projects) + 1)
    unit_end = max(2, len(units) + 1)
    project_validation = DataValidation(
        type="list",
        formula1=f"Lists!$A$2:$A${project_end}",
        allow_blank=False,
    )
    unit_validation = DataValidation(
        type="list",
        formula1=f"Lists!$B$2:$B${unit_end}",
        allow_blank=False,
    )
    sheet.add_data_validation(project_validation)
    sheet.add_data_validation(unit_validation)
    project_validation.add("A2:A1000")
    unit_validation.add("G2:G1000")
    lists.sheet_state = "hidden"

    output = BytesIO()
    workbook.save(output)
    ExportAudit.objects.create(
        dataset=ExportAudit.Dataset.OPENING_TEMPLATE,
        file_format=ExportAudit.Format.XLSX,
        filters={},
        columns=headers,
        row_count=0,
        created_by=user,
    )
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="opening-stock-template.xlsx"'
    response["Cache-Control"] = "private, no-store"
    return response
