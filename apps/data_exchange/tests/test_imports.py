import io
import tempfile
import uuid
from datetime import date
from decimal import Decimal
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from apps.inventory.models import StockItem, StockMovement, Unit
from apps.projects.models import Project

from ..models import ImportJob, ImportRow
from ..services.importing import (
    ImportProcessingError,
    confirm_import,
    preview_import,
)

User = get_user_model()


def workbook_file(name, sheet_name, headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    payload = io.BytesIO()
    workbook.save(payload)
    return SimpleUploadedFile(
        name,
        payload.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class ImportWorkflowTests(TestCase):
    def setUp(self):
        self.media = tempfile.TemporaryDirectory()
        self.addCleanup(self.media.cleanup)
        self.override = override_settings(MEDIA_ROOT=self.media.name)
        self.override.enable()
        self.addCleanup(self.override.disable)
        self.admin = User.objects.create_superuser(
            username="admin",
            email="admin@example.com",
            password="safe-password",
        )
        self.keeper = User.objects.create_user(username="keeper", password="safe-password")
        self.project = Project.objects.create(code="ARAMCO-01", name="Aramco Project")
        self.unit = Unit.objects.get(normalized_name="bag")

    def legacy_job(self, rows, *, update_existing=True):
        file = workbook_file(
            "legacy.xlsx",
            "Database",
            [
                "Material Name",
                "Description",
                "Name of Supplier",
                "Contact / Mobile No.",
                "Location of Supplier",
                "Unit Price",
                "Last Purchase Date",
            ],
            rows,
        )
        return ImportJob.objects.create(
            import_type=ImportJob.Type.LEGACY_CATALOG,
            source_file=file,
            original_filename=file.name,
            project=self.project,
            default_unit=self.unit,
            options={"update_existing_records": update_existing},
            created_by=self.admin,
        )

    def opening_job(self, rows):
        file = workbook_file(
            "opening.xlsx",
            "Opening Stock",
            [
                "Project Code",
                "Material Name",
                "Description",
                "Supplier Name",
                "Supplier Phone",
                "Supplier Location",
                "Unit",
                "Opening Quantity",
                "Unit Price",
                "Opening Date",
                "Minimum Quantity",
                "Reference",
                "Notes",
            ],
            rows,
        )
        return ImportJob.objects.create(
            import_type=ImportJob.Type.OPENING_STOCK,
            source_file=file,
            original_filename=file.name,
            created_by=self.admin,
        )

    def test_legacy_import_creates_zero_balance_record_and_preserves_snapshot(self):
        job = self.legacy_job(
            [[
                "Cement",
                "Portland 40 KG",
                "Gulf Cement",
                573686575,
                "Dammam",
                50,
                date(2026, 7, 30),
            ]]
        )
        preview_import(job)
        job.refresh_from_db()
        self.assertEqual(job.total_rows, 1)
        row = job.rows.get()
        self.assertEqual(row.planned_action, ImportRow.Action.CREATE)
        self.assertIn("lost a leading zero", row.message)
        confirm_import(job=job, user=self.admin)
        item = StockItem.objects.get()
        self.assertEqual(item.current_quantity, Decimal("0"))
        self.assertEqual(item.latest_unit_price, Decimal("50"))
        self.assertEqual(item.latest_addition_date, date(2026, 7, 30))
        self.assertEqual(StockMovement.objects.count(), 0)

    def test_legacy_exact_match_updates_without_changing_quantity(self):
        item = StockItem.objects.create(
            project=self.project,
            material_name="Cement",
            supplier_name="Gulf Cement",
            supplier_phone="573686575",
            unit=self.unit,
        )
        job = self.legacy_job(
            [[
                "Cement",
                "Updated description",
                "Gulf Cement",
                "573686575",
                "Riyadh",
                55,
                date(2026, 8, 1),
            ]]
        )
        preview_import(job)
        self.assertEqual(job.rows.get().planned_action, ImportRow.Action.UPDATE)
        confirm_import(job=job, user=self.admin)
        item.refresh_from_db()
        self.assertEqual(item.current_quantity, Decimal("0"))
        self.assertEqual(item.description, "Updated description")
        self.assertEqual(item.latest_unit_price, Decimal("55"))

    def test_duplicate_legacy_identity_in_same_workbook_is_skipped(self):
        row = ["Cement", "", "Gulf Cement", "573686575", "Dammam", 50, date(2026, 7, 30)]
        job = self.legacy_job([row, row])
        preview_import(job)
        self.assertEqual(job.rows.filter(planned_action=ImportRow.Action.SKIP).count(), 1)
        confirm_import(job=job, user=self.admin)
        self.assertEqual(StockItem.objects.count(), 1)

    def test_opening_import_creates_immutable_opening_movement(self):
        job = self.opening_job(
            [[
                self.project.code,
                "Cement",
                "Portland 40 KG",
                "Gulf Cement",
                "+966 50 111 1111",
                "Dammam",
                self.unit.symbol,
                50,
                24.5,
                date(2026, 8, 1),
                10,
                "OPEN-1",
                "Initial count",
            ]]
        )
        preview_import(job)
        confirm_import(job=job, user=self.admin)
        item = StockItem.objects.get()
        movement = StockMovement.objects.get()
        self.assertEqual(item.current_quantity, Decimal("50"))
        self.assertEqual(movement.movement_type, StockMovement.Type.OPENING)
        self.assertEqual(movement.invoice_reference, "OPEN-1")
        self.assertEqual(movement.previous_balance, Decimal("0"))

    def test_opening_import_updates_empty_matching_record_metadata(self):
        item = StockItem.objects.create(
            project=self.project,
            material_name="Cement",
            description="Old description",
            supplier_name="Gulf Cement",
            supplier_phone="+966 50 111 1111",
            supplier_location="Old location",
            unit=self.unit,
            minimum_quantity=Decimal("1"),
        )
        job = self.opening_job(
            [[
                self.project.code,
                "Cement",
                "Portland 40 KG",
                "Gulf Cement",
                "+966 50 111 1111",
                "Dammam",
                self.unit.symbol,
                50,
                24.5,
                date(2026, 8, 1),
                10,
                "OPEN-1",
                "Initial count",
            ]]
        )
        preview_import(job)
        self.assertEqual(job.rows.get().planned_action, ImportRow.Action.OPEN_EXISTING)
        confirm_import(job=job, user=self.admin)
        item.refresh_from_db()
        self.assertEqual(item.description, "Portland 40 KG")
        self.assertEqual(item.supplier_location, "Dammam")
        self.assertEqual(item.minimum_quantity, Decimal("10"))
        self.assertEqual(item.current_quantity, Decimal("50"))

    def test_opening_import_rejects_existing_movement_history(self):
        from apps.inventory.services.stock import add_stock

        add_stock(
            user=self.keeper,
            idempotency_key=uuid.uuid4(),
            quantity=Decimal("1"),
            movement_date=date(2026, 8, 1),
            project=self.project,
            material_name="Cement",
            supplier_name="Gulf Cement",
            supplier_phone="+966 50 111 1111",
            unit=self.unit,
        )
        job = self.opening_job(
            [[
                self.project.code,
                "Cement",
                "",
                "Gulf Cement",
                "+966 50 111 1111",
                "",
                self.unit.symbol,
                50,
                24.5,
                date(2026, 8, 1),
                0,
                "",
                "",
            ]]
        )
        preview_import(job)
        self.assertEqual(job.rows.get().status, ImportRow.Status.ERROR)

    def test_confirmation_failure_rolls_back_every_row(self):
        job = self.opening_job(
            [
                [
                    self.project.code, "Cement A", "", "Supplier", "501", "",
                    self.unit.symbol, 10, 1, date(2026, 8, 1), 0, "", "",
                ],
                [
                    self.project.code, "Cement B", "", "Supplier", "502", "",
                    self.unit.symbol, 10, 1, date(2026, 8, 1), 0, "", "",
                ],
            ]
        )
        preview_import(job)
        original = StockItem.save
        calls = {"count": 0}

        def fail_second(instance, *args, **kwargs):
            calls["count"] += 1
            if calls["count"] == 2:
                raise RuntimeError("simulated failure")
            return original(instance, *args, **kwargs)

        with patch.object(StockItem, "save", fail_second):
            with self.assertRaises(ImportProcessingError):
                confirm_import(job=job, user=self.admin)
        self.assertEqual(StockItem.objects.count(), 0)
        self.assertEqual(StockMovement.objects.count(), 0)

    def test_opening_template_contains_validated_project_and_unit_lists(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("data_exchange:opening_template"))
        generated = load_workbook(io.BytesIO(response.content), data_only=False)
        self.assertEqual(
            generated.sheetnames,
            ["Instructions", "Opening Stock", "Lists"],
        )
        validations = generated["Opening Stock"].data_validations.dataValidation
        self.assertEqual(len(validations), 2)
        self.assertEqual(generated["Lists"]["A2"].value, self.project.code)
        self.assertEqual(generated["Lists"].sheet_state, "hidden")
        self.assertEqual(generated["Instructions"]["A8"].value, "Accepted columns")
        self.assertEqual(generated["Instructions"]["A9"].value, "Project Code")
        self.assertEqual(generated["Opening Stock"]["E2"].number_format, "@")
        self.assertEqual(generated["Opening Stock"]["J2"].number_format, "yyyy-mm-dd")
        generated.close()

    def test_opening_upload_page_explains_accepted_columns(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("data_exchange:opening_import"))
        self.assertContains(response, "Accepted opening-stock workbook format")
        self.assertContains(response, "Project Code")
        self.assertContains(response, "Supplier Phone")
        self.assertContains(response, "YYYY-MM-DD recommended")
        self.assertContains(response, "Maximum file size: 20 MB")

    def test_import_pages_require_administrator(self):
        self.client.force_login(self.keeper)
        self.assertEqual(self.client.get(reverse("data_exchange:import_list")).status_code, 403)
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(reverse("data_exchange:import_list")).status_code, 200)
        self.assertEqual(
            self.client.get(reverse("data_exchange:opening_template")).status_code,
            200,
        )
