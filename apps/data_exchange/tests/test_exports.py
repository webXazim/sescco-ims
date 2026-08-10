import csv
import io
import uuid
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.inventory.models import Unit
from apps.inventory.services.stock import add_stock, use_stock
from apps.projects.models import Project

from ..models import ExportAudit

User = get_user_model()


class FilteredExportTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="keeper", password="safe-password")
        self.client.force_login(self.user)
        self.project = Project.objects.create(code="ARAMCO-01", name="Aramco Project")
        self.other_project = Project.objects.create(code="NEOM-02", name="NEOM Project")
        self.unit = Unit.objects.get(normalized_name="bag")
        self.today = timezone.localdate()

    def add(self, *, index=1, project=None, material=None):
        return add_stock(
            user=self.user,
            idempotency_key=uuid.uuid4(),
            quantity=Decimal("10"),
            movement_date=self.today,
            project=project or self.project,
            material_name=material or f"Cement {index}",
            description="50 KG construction material",
            supplier_name="Gulf Cement",
            supplier_phone=f"+966 50 100 {index:04d}",
            supplier_location="Dammam",
            unit=self.unit,
            minimum_quantity=Decimal("2"),
            unit_price=Decimal("24.50"),
            invoice_reference=f"INV-{index}",
        ).movement.stock_item

    def test_csv_exports_every_filtered_row_not_only_current_page(self):
        for index in range(1, 56):
            self.add(index=index)
        response = self.client.get(
            reverse("data_exchange:inventory_export", args=["csv"]),
            {"q": "cement", "columns": ["project", "material", "quantity"]},
        )
        self.assertEqual(response.status_code, 200)
        text = response.content.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(text)))
        self.assertEqual(rows[0], ["Project", "Material", "Quantity"])
        self.assertEqual(len(rows) - 1, 55)
        audit = ExportAudit.objects.get()
        self.assertEqual(audit.row_count, 55)
        self.assertEqual(audit.columns, ["project", "material", "quantity"])

    def test_xlsx_contains_results_and_export_information(self):
        self.add(index=1)
        response = self.client.get(
            reverse("data_exchange:inventory_export", args=["xlsx"]),
            {"project": self.project.code, "columns": ["material", "price"]},
        )
        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
        self.assertEqual(workbook.sheetnames, ["Results", "Export Information"])
        self.assertEqual(workbook["Results"]["A1"].value, "Material")
        self.assertEqual(workbook["Results"]["B2"].value, 24.5)
        metadata = {
            row[0].value: row[1].value
            for row in workbook["Export Information"].iter_rows(min_row=2)
            if row[0].value
        }
        self.assertEqual(metadata["Matching rows"], 1)
        self.assertIn("Material", metadata["Columns"])

    def test_spreadsheet_formula_injection_is_neutralized(self):
        self.add(index=1, material='  =HYPERLINK("https://example.test")')
        response = self.client.get(
            reverse("data_exchange:inventory_export", args=["csv"]),
            {"columns": ["material"]},
        )
        rows = list(csv.reader(io.StringIO(response.content.decode("utf-8-sig"))))
        self.assertTrue(rows[1][0].lstrip("'").lstrip().startswith("="))
        self.assertTrue(rows[1][0].startswith("'"))

    def test_activity_export_uses_current_filters(self):
        item = self.add(index=1)
        use_stock(
            stock_item=item,
            user=self.user,
            idempotency_key=uuid.uuid4(),
            quantity=Decimal("2"),
            movement_date=self.today,
            purpose="Block work",
        )
        response = self.client.get(
            reverse("data_exchange:activity_export", args=["csv"]),
            {"movement_type": "usage", "columns": ["type", "material", "quantity"]},
        )
        rows = list(csv.reader(io.StringIO(response.content.decode("utf-8-sig"))))
        self.assertEqual(len(rows) - 1, 1)
        self.assertEqual(rows[1][0], "Stock used")
        self.assertEqual(rows[1][1], "Cement 1")

    def test_stock_history_export_never_leaks_another_stock_record(self):
        first = self.add(index=1)
        self.add(index=2)
        response = self.client.get(
            reverse(
                "data_exchange:stock_history_export",
                args=[first.reference, "csv"],
            )
        )
        text = response.content.decode("utf-8-sig")
        self.assertIn("Cement 1", text)
        self.assertNotIn("Cement 2", text)

    def test_project_inventory_export_stays_in_selected_project(self):
        self.add(index=1, project=self.project)
        self.add(index=2, project=self.other_project)
        response = self.client.get(
            reverse(
                "data_exchange:project_inventory_export",
                args=[self.project.code, "csv"],
            )
        )
        text = response.content.decode("utf-8-sig")
        self.assertIn("Cement 1", text)
        self.assertNotIn("Cement 2", text)
